
from __future__ import annotations

import argparse
import json
import math
import os
import re
import struct
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    import pandas as pd
except Exception:
    pd = None

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None


MAGIC = b"UIKenn0File"

# Validated against the user's PVPM1540X datasets.
OFFSETS = {
    "format_version": 153,
    "device_model_serial": 160,
    "device_calibration_date": 175,
    "sensor_name": 233,
    "sensor_type": 247,
    "sensor_calibration_date": 280,
    "module_manufacturer": 300,
    "module_part_number": 326,
    "module_technology": 375,
    "label_date_short": 2477,
    "label_time": 2486,
    "label_site_name": 2546,
    "label_string_name": 2597,
    "label_timestamp_text": 2648,
    "module_imp_stc": 388,
    "module_voc_stc": 392,
    "module_isc_stc": 396,
    "module_vmp_stc": 400,
    "module_pmax_stc": 404,
    "module_temp_coeff_current": 416,
    "module_temp_coeff_voltage": 420,
    "module_temp_coeff_power": 424,
    "curve_current_start": 503,
    "curve_current_count": 110,
    "curve_voltage_start": 1475,
    "curve_voltage_count": 130,
}

XLS_COLUMN_MAP = {
    "Dateiname": "filename",
    "Filename": "filename",
    "Datum Zeit": "datetime_text",
    "PVPM Nr.": "device_number",
    "Einstr.-Sensor Nr.:": "sensor_number",
    "T sens": "t_sens",
    "T mod": "t_mod",
    "E eff": "e_eff",
    "Isc": "isc",
    "Uoc": "uoc",
    "Ipmax": "ipmax",
    "Upmax": "upmax",
    "Pmax": "pmax",
    "Isc 0": "isc_0",
    "Uoc 0": "uoc_0",
    "Ipmax0": "ipmax_0",
    "Upmax0": "upmax_0",
    "Ppk": "ppk",
    "Fill factor": "fill_factor",
    "Rs": "rs",
    "Rp": "rp",
    "Module": "module",
    "Module.1": "module_1",
    "Module.2": "module_2",
    "Module.3": "module_3",
    "Module.4": "module_4",
    "Customer": "customer",
    "Plant": "plant",
    "Part": "part",
}

SEMANTIC_WARNING = (
    "SUI label fields are treated as raw labels only. "
    "Plant/Part/module/numeric metrics should prefer XLS when available."
)


class SUIParseError(Exception):
    pass


@dataclass
class ArchiveEntry:
    name: str
    data: bytes


def normalize_name(name: str | Path) -> str:
    return os.path.basename(str(name)).strip().lower()


def strip_ext(name: str | Path) -> str:
    return os.path.splitext(os.path.basename(str(name)))[0].strip().lower()


def f32(data: bytes, offset: int) -> float | None:
    if offset + 4 > len(data):
        return None
    try:
        return struct.unpack_from("<f", data, offset)[0]
    except Exception:
        return None


def clean_text(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.replace("\x00", "").strip()
    return value or None


def read_lp_string(data: bytes, offset: int, max_len: int = 120) -> str | None:
    i = offset
    while i < len(data) and data[i] == 0:
        i += 1
    if i >= len(data):
        return None
    length = data[i]
    if length <= 0 or length > max_len or i + 1 + length > len(data):
        return None
    raw = data[i + 1 : i + 1 + length]
    try:
        text = raw.decode("latin1", errors="ignore")
    except Exception:
        return None
    return clean_text(text)


def ascii_strings(data: bytes, min_len: int = 4) -> list[str]:
    found = re.findall(rb"[ -~]{%d,}" % min_len, data)
    return [s.decode("latin1", errors="ignore") for s in found]


def extract_float_series(data: bytes, start: int, count: int) -> list[float]:
    values: list[float] = []
    for i in range(count):
        off = start + i * 4
        if off + 4 > len(data):
            break
        val = f32(data, off)
        if val is None:
            break
        values.append(val)
    return values


def trim_tail(values: list[float], eps: float = 1e-12) -> list[float]:
    last = -1
    for i, v in enumerate(values):
        if math.isfinite(v) and abs(v) > eps:
            last = i
    return values[: last + 1] if last >= 0 else []


def curve_points_from_blocks(data: bytes) -> list[dict[str, float]]:
    current = trim_tail(
        extract_float_series(data, OFFSETS["curve_current_start"], OFFSETS["curve_current_count"])
    )
    voltage = trim_tail(
        extract_float_series(data, OFFSETS["curve_voltage_start"], OFFSETS["curve_voltage_count"])
    )

    if not current or not voltage:
        return []

    while len(voltage) > 3 and voltage[0] > voltage[1] + 1e-6:
        voltage = voltage[1:]

    n = min(len(current), len(voltage))
    points: list[dict[str, float]] = []
    for i in range(n):
        u = float(voltage[i])
        ia = float(current[i])
        if not (math.isfinite(u) and math.isfinite(ia)):
            continue
        if u < -1 or ia < -1:
            continue
        if u > 2500 or ia > 100:
            continue
        points.append({"u_v": u, "i_a": ia, "p_w": u * ia})
    return points


def derive_metrics_from_curve(points: list[dict[str, float]]) -> dict[str, Any]:
    if not points:
        return {"confidence": "none"}

    points_sorted = sorted(points, key=lambda p: p["u_v"])
    isc = max(points_sorted, key=lambda p: p["i_a"])["i_a"]
    voc = max(points_sorted, key=lambda p: p["u_v"])["u_v"]
    mpp = max(points_sorted, key=lambda p: p["p_w"])
    imp = mpp["i_a"]
    vmp = mpp["u_v"]
    pmax = mpp["p_w"]
    ff = (pmax / (voc * isc)) if voc and isc else None

    return {
        "isc_a": round(isc, 6),
        "voc_v": round(voc, 6),
        "imp_a": round(imp, 6),
        "vmp_v": round(vmp, 6),
        "pmax_w": round(pmax, 6),
        "fill_factor": round(ff, 6) if ff is not None else None,
        "point_count": len(points_sorted),
        "confidence": "medium",
    }


def read_excel_to_df(path_or_buffer: Any) -> "pd.DataFrame":
    """
    Read legacy XLS as robustly as possible.
    Requires pandas and usually xlrd for .xls support.
    """
    if pd is None:
        raise RuntimeError("pandas is not installed")

    last_err = None
    for engine in (None, "xlrd", "openpyxl"):
        try:
            kwargs = {}
            if engine is not None:
                kwargs["engine"] = engine
            return pd.read_excel(path_or_buffer, **kwargs)
        except Exception as e:
            last_err = e
    raise RuntimeError(
        "Failed to read Excel. For .xls sidecars, install xlrd in the runtime environment. "
        f"Last error: {last_err}"
    )


def normalize_xls_columns(columns: list[Any]) -> list[str]:
    out: list[str] = []
    for c in columns:
        c_str = str(c).strip()
        out.append(XLS_COLUMN_MAP.get(c_str, c_str))
    return out


def read_master_xls(path: str | Path) -> dict[str, dict[str, Any]]:
    df = read_excel_to_df(path)
    df.columns = normalize_xls_columns(list(df.columns))

    if "filename" not in df.columns:
        raise RuntimeError("Could not find filename column in XLS")

    df = df[df["filename"].notna()].copy()
    df["filename"] = df["filename"].astype(str).str.strip()
    df = df[df["filename"].str.lower() != "filename"].copy()

    records: dict[str, dict[str, Any]] = {}
    for _, row in df.iterrows():
        record: dict[str, Any] = {}
        for col in df.columns:
            val = row[col]
            if pd.isna(val):
                continue
            if hasattr(val, "item"):
                try:
                    val = val.item()
                except Exception:
                    pass
            record[col] = val
        records[normalize_name(record["filename"])] = record
    return records


def read_sidecar_xls_bytes(data: bytes, filename: str) -> dict[str, Any]:
    """
    Sidecar XLS often contains the detailed export for a single measurement.
    This parser intentionally stays forgiving:
    - top key/value rows -> metadata
    - if curve headers appear, curve parsing is left to SUI
    """
    df = read_excel_to_df(__import__("io").BytesIO(data))
    top_values: dict[str, Any] = {}

    # 2-column style rows at the top
    for _, row in df.head(50).iterrows():
        vals = [v for v in row.tolist() if not (pd.isna(v) if pd is not None else False)]
        vals = [str(v).strip() for v in vals if str(v).strip()]
        if len(vals) >= 2 and len(vals[0]) <= 40:
            top_values[vals[0]] = vals[1]

    return {
        "filename": filename,
        "top_values": top_values,
    }


def parse_sui_bytes(
    data: bytes,
    file_name: str,
    master_row: dict[str, Any] | None = None,
    sidecar_xls: dict[str, Any] | None = None,
    include_curve: bool = True,
    include_ascii: bool = False,
) -> dict[str, Any]:
    if not data.startswith(MAGIC):
        raise SUIParseError(f"{file_name}: unrecognized header")

    result: dict[str, Any] = {
        "file_name": file_name,
        "file_size": len(data),
        "format": "UIKenn0File",
        "parser_version": "v4.1",
        "variant": {
            "family": "PVPMdisp_4_302_fixed_2957",
        },
        "semantic_warning": SEMANTIC_WARNING,
        "format_version": read_lp_string(data, OFFSETS["format_version"], 20),
        "raw_sui_labels": {
            "site_label_raw": read_lp_string(data, OFFSETS["label_site_name"], 80),
            "string_label_raw": read_lp_string(data, OFFSETS["label_string_name"], 80),
        },
        "device": {
            "model_serial": read_lp_string(data, OFFSETS["device_model_serial"], 40),
            "calibration_date": read_lp_string(data, OFFSETS["device_calibration_date"], 20),
        },
        "sensor": {
            "name": read_lp_string(data, OFFSETS["sensor_name"], 60),
            "type": read_lp_string(data, OFFSETS["sensor_type"], 20),
            "calibration_date": read_lp_string(data, OFFSETS["sensor_calibration_date"], 20),
        },
        "module_reference_from_sui": {
            "manufacturer": read_lp_string(data, OFFSETS["module_manufacturer"], 40),
            "part_number_raw": read_lp_string(data, OFFSETS["module_part_number"], 80),
            "technology": read_lp_string(data, OFFSETS["module_technology"], 20),
            "imp_stc": f32(data, OFFSETS["module_imp_stc"]),
            "voc_stc": f32(data, OFFSETS["module_voc_stc"]),
            "isc_stc": f32(data, OFFSETS["module_isc_stc"]),
            "vmp_stc": f32(data, OFFSETS["module_vmp_stc"]),
            "pmax_stc": f32(data, OFFSETS["module_pmax_stc"]),
            "temp_coeff_current": f32(data, OFFSETS["module_temp_coeff_current"]),
            "temp_coeff_voltage": f32(data, OFFSETS["module_temp_coeff_voltage"]),
            "temp_coeff_power": f32(data, OFFSETS["module_temp_coeff_power"]),
        },
        "measurement_from_sui": {
            "date_short": read_lp_string(data, OFFSETS["label_date_short"], 20),
            "time": read_lp_string(data, OFFSETS["label_time"], 20),
            "timestamp_text": read_lp_string(data, OFFSETS["label_timestamp_text"], 40),
        },
    }

    if include_curve:
        points = curve_points_from_blocks(data)
        result["curve"] = {
            "point_count": len(points),
            "points": points,
        }
        result["derived_from_curve"] = derive_metrics_from_curve(points)

    if include_ascii:
        result["ascii_strings"] = ascii_strings(data)

    if sidecar_xls is not None:
        result["sidecar_xls"] = sidecar_xls

    if master_row is not None:
        result["validated_xls_labels"] = {
            "plant": master_row.get("plant"),
            "part": master_row.get("part"),
            "module": master_row.get("module"),
            "customer": master_row.get("customer"),
            "datetime_text": master_row.get("datetime_text"),
            "device_number": master_row.get("device_number"),
            "sensor_number": master_row.get("sensor_number"),
        }
        result["validated_metrics"] = {
            "t_sens": master_row.get("t_sens"),
            "t_mod": master_row.get("t_mod"),
            "e_eff": master_row.get("e_eff"),
            "isc": master_row.get("isc"),
            "uoc": master_row.get("uoc"),
            "ipmax": master_row.get("ipmax"),
            "upmax": master_row.get("upmax"),
            "pmax": master_row.get("pmax"),
            "isc_0": master_row.get("isc_0"),
            "uoc_0": master_row.get("uoc_0"),
            "ipmax_0": master_row.get("ipmax_0"),
            "upmax_0": master_row.get("upmax_0"),
            "ppk": master_row.get("ppk"),
            "fill_factor": master_row.get("fill_factor"),
            "rs": master_row.get("rs"),
            "rp": master_row.get("rp"),
        }
        result["normalized"] = {
            "site_name": master_row.get("plant"),
            "part_name": master_row.get("part"),
            "module_part_number": master_row.get("module"),
            "timestamp": master_row.get("datetime_text"),
        }
        result["confidence"] = {
            "metadata_from_sui": "high",
            "curve_from_sui": "medium",
            "metrics_from_master_xls": "high",
            "overall": "high",
        }
    else:
        result["normalized"] = {
            "site_name": result["raw_sui_labels"]["site_label_raw"],
            "part_name": result["raw_sui_labels"]["string_label_raw"],
            "module_part_number": result["module_reference_from_sui"]["part_number_raw"],
            "timestamp": result["measurement_from_sui"]["timestamp_text"],
        }
        result["confidence"] = {
            "metadata_from_sui": "high",
            "curve_from_sui": "medium",
            "metrics_from_master_xls": "none",
            "overall": "medium",
        }

    return result


def parse_zip_dataset(
    zip_path: str | Path,
    master_xls: str | Path | None = None,
    include_curve: bool = True,
    include_ascii: bool = False,
) -> list[dict[str, Any]]:
    master_index = read_master_xls(master_xls) if master_xls else None

    with zipfile.ZipFile(zip_path, "r") as zf:
        name_map = {normalize_name(n): n for n in zf.namelist()}
        stem_to_sidecar_xls = {
            strip_ext(n): n
            for n in zf.namelist()
            if n.lower().endswith(".xls") and normalize_name(n) != "all.xls"
        }
        sui_names = [n for n in zf.namelist() if n.lower().endswith(".sui")]

        results: list[dict[str, Any]] = []
        for sui_name in sorted(sui_names):
            sui_bytes = zf.read(sui_name)
            file_name_only = os.path.basename(sui_name)
            stem = strip_ext(file_name_only)

            master_row = None
            if master_index:
                master_row = master_index.get(normalize_name(file_name_only))

            sidecar = None
            if stem in stem_to_sidecar_xls:
                xls_name = stem_to_sidecar_xls[stem]
                try:
                    sidecar = read_sidecar_xls_bytes(zf.read(xls_name), os.path.basename(xls_name))
                except Exception as e:
                    sidecar = {"filename": os.path.basename(xls_name), "warning": str(e)}

            results.append(
                parse_sui_bytes(
                    sui_bytes,
                    file_name_only,
                    master_row=master_row,
                    sidecar_xls=sidecar,
                    include_curve=include_curve,
                    include_ascii=include_ascii,
                )
            )
    return results


def parse_single(
    sui_path: str | Path,
    master_xls: str | Path | None = None,
    include_curve: bool = True,
    include_ascii: bool = False,
) -> dict[str, Any]:
    master_index = read_master_xls(master_xls) if master_xls else None
    path = Path(sui_path)
    master_row = master_index.get(normalize_name(path.name)) if master_index else None
    return parse_sui_bytes(
        path.read_bytes(),
        path.name,
        master_row=master_row,
        sidecar_xls=None,
        include_curve=include_curve,
        include_ascii=include_ascii,
    )


def write_single_excel(parsed: dict[str, Any], output_path: str | Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["field", "value"])

    rows = [
        ("file_name", parsed.get("file_name")),
        ("format_version", parsed.get("format_version")),
        ("device_model_serial", parsed.get("device", {}).get("model_serial")),
        ("sensor_name", parsed.get("sensor", {}).get("name")),
        ("site_name_normalized", parsed.get("normalized", {}).get("site_name")),
        ("part_name_normalized", parsed.get("normalized", {}).get("part_name")),
        ("module_part_number_normalized", parsed.get("normalized", {}).get("module_part_number")),
        ("timestamp_normalized", parsed.get("normalized", {}).get("timestamp")),
        ("site_label_raw", parsed.get("raw_sui_labels", {}).get("site_label_raw")),
        ("string_label_raw", parsed.get("raw_sui_labels", {}).get("string_label_raw")),
    ]
    for k, v in parsed.get("validated_metrics", {}).items():
        rows.append((f"validated_{k}", v))
    for k, v in parsed.get("derived_from_curve", {}).items():
        rows.append((f"derived_{k}", v))
    for row in rows:
        ws.append(list(row))

    ws2 = wb.create_sheet("Curve")
    ws2.append(["u_v", "i_a", "p_w"])
    for p in parsed.get("curve", {}).get("points", []):
        ws2.append([p.get("u_v"), p.get("i_a"), p.get("p_w")])

    ws3 = wb.create_sheet("Raw_JSON")
    ws3["A1"] = json.dumps(parsed, ensure_ascii=False, indent=2)
    wb.save(output_path)


def write_batch_excel(items: list[dict[str, Any]], output_path: str | Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Batch"

    headers = [
        "file_name",
        "device_model_serial",
        "site_name_normalized",
        "part_name_normalized",
        "module_part_number_normalized",
        "site_label_raw",
        "string_label_raw",
        "timestamp_normalized",
        "isc",
        "uoc",
        "ipmax",
        "upmax",
        "pmax",
        "fill_factor",
        "rs",
        "rp",
        "derived_pmax_w",
        "curve_point_count",
        "overall_confidence",
    ]
    ws.append(headers)

    for item in items:
        ws.append([
            item.get("file_name"),
            item.get("device", {}).get("model_serial"),
            item.get("normalized", {}).get("site_name"),
            item.get("normalized", {}).get("part_name"),
            item.get("normalized", {}).get("module_part_number"),
            item.get("raw_sui_labels", {}).get("site_label_raw"),
            item.get("raw_sui_labels", {}).get("string_label_raw"),
            item.get("normalized", {}).get("timestamp"),
            item.get("validated_metrics", {}).get("isc"),
            item.get("validated_metrics", {}).get("uoc"),
            item.get("validated_metrics", {}).get("ipmax"),
            item.get("validated_metrics", {}).get("upmax"),
            item.get("validated_metrics", {}).get("pmax"),
            item.get("validated_metrics", {}).get("fill_factor"),
            item.get("validated_metrics", {}).get("rs"),
            item.get("validated_metrics", {}).get("rp"),
            item.get("derived_from_curve", {}).get("pmax_w"),
            item.get("curve", {}).get("point_count"),
            item.get("confidence", {}).get("overall"),
        ])
    wb.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="PVPM SUI parser v4.1 with ZIP sidecar-XLS reconciliation and raw/validated split"
    )
    p.add_argument("input", help="A .SUI file or a ZIP containing .SUI/.XLS pairs")
    p.add_argument("--master-xls", help="Optional master XLS like hamadya.xls / all.XLS")
    p.add_argument("--batch", action="store_true", help="Required for ZIP batch processing")
    p.add_argument("--output-format", choices=["json", "excel"], default="json")
    p.add_argument("-o", "--output", help="Output path")
    p.add_argument("--ascii", action="store_true", help="Include printable ASCII strings from SUI")
    p.add_argument("--no-curve", action="store_true", help="Skip SUI curve extraction")
    return p


def main() -> None:
    args = build_parser().parse_args()
    include_curve = not args.no_curve
    input_path = Path(args.input)

    if input_path.suffix.lower() == ".zip":
        items = parse_zip_dataset(
            input_path,
            master_xls=args.master_xls,
            include_curve=include_curve,
            include_ascii=args.ascii,
        )
        if args.output_format == "json":
            text = json.dumps(items, ensure_ascii=False, indent=2)
            if args.output:
                Path(args.output).write_text(text, encoding="utf-8")
            else:
                print(text)
        else:
            out = args.output or "pvpm_batch_v4_1.xlsx"
            write_batch_excel(items, out)
            print(out)
        return

    parsed = parse_single(
        input_path,
        master_xls=args.master_xls,
        include_curve=include_curve,
        include_ascii=args.ascii,
    )
    if args.output_format == "json":
        text = json.dumps(parsed, ensure_ascii=False, indent=2)
        if args.output:
            Path(args.output).write_text(text, encoding="utf-8")
        else:
            print(text)
    else:
        out = args.output or f"{input_path.stem}.parsed.v4_1.xlsx"
        write_single_excel(parsed, out)
        print(out)


if __name__ == "__main__":
    main()
