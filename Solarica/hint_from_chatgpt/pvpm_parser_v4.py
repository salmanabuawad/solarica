
from __future__ import annotations

import argparse
import json
import math
import os
import re
import struct
import zipfile
from pathlib import Path
from typing import Any

try:
    import pandas as pd
except Exception:
    pd = None

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None


MAGIC = b"UIKenn0File"

# Validated primarily against the Hamadya dataset family:
# - fixed-size SUI files (~2957 bytes)
# - format_version "4.302"
# - device family PVPM1540X
OFFSETS = {
    # length-prefixed strings
    "format_version": 153,
    "device_model_serial": 160,
    "device_calibration_date": 175,
    "sensor_name": 233,
    "sensor_type": 247,
    "sensor_calibration_date": 280,
    "module_manufacturer": 300,
    "module_part_number": 326,
    "module_technology": 375,
    "label_date_short": 2477,
    "label_time": 2486,
    "label_site_name": 2546,
    "label_string_name": 2597,
    "label_timestamp_text": 2648,
    # reference/module float block
    "module_imp_stc": 388,
    "module_voc_stc": 392,
    "module_isc_stc": 396,
    "module_vmp_stc": 400,
    "module_pmax_stc": 404,
    "module_temp_coeff_current": 416,
    "module_temp_coeff_voltage": 420,
    "module_temp_coeff_power": 424,
    # curve blocks
    "curve_current_start": 503,
    "curve_current_count": 110,
    "curve_voltage_start": 1475,
    "curve_voltage_count": 130,
}

XLS_COLUMN_MAP = {
    "Dateiname": "filename",
    "Filename": "filename",
    "Dateiname ": "filename",
    "Datum Zeit": "datetime_text",
    "PVPM Nr.": "device_number",
    "Einstr.-Sensor Nr.:": "sensor_number",
    "T sens": "t_sens",
    "T mod": "t_mod",
    "E eff": "e_eff",
    "Isc": "isc",
    "Uoc": "uoc",
    "Ipmax": "ipmax",
    "Upmax": "upmax",
    "Pmax": "pmax",
    "Isc 0": "isc_0",
    "Uoc 0": "uoc_0",
    "Ipmax0": "ipmax_0",
    "Upmax0": "upmax_0",
    "Ppk": "ppk",
    "Fill factor": "fill_factor",
    "Rs": "rs",
    "Rp": "rp",
    "Module": "module",
    "Module.1": "module_1",
    "Module.2": "module_2",
    "Module.3": "module_3",
    "Module.4": "module_4",
    "Customer": "customer",
    "Plant": "plant",
    "Part": "part",
}


class SUIParseError(Exception):
    pass


def f32(data: bytes, offset: int) -> float | None:
    if offset + 4 > len(data):
        return None
    try:
        return struct.unpack_from("<f", data, offset)[0]
    except Exception:
        return None


def clean_text(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.replace("\x00", "").strip()
    return value or None


def read_lp_string(data: bytes, offset: int, max_len: int = 120) -> str | None:
    i = offset
    while i < len(data) and data[i] == 0:
        i += 1
    if i >= len(data):
        return None
    length = data[i]
    if length <= 0 or length > max_len or i + 1 + length > len(data):
        return None
    raw = data[i + 1 : i + 1 + length]
    try:
        text = raw.decode("latin1", errors="ignore")
    except Exception:
        return None
    return clean_text(text)


def ascii_strings(data: bytes, min_len: int = 4) -> list[str]:
    found = re.findall(rb"[ -~]{%d,}" % min_len, data)
    return [s.decode("latin1", errors="ignore") for s in found]


def extract_float_series(data: bytes, start: int, count: int) -> list[float]:
    values: list[float] = []
    for i in range(count):
        off = start + i * 4
        if off + 4 > len(data):
            break
        val = f32(data, off)
        if val is None:
            break
        values.append(val)
    return values


def trim_tail(values: list[float], eps: float = 1e-12) -> list[float]:
    last = -1
    for i, v in enumerate(values):
        if math.isfinite(v) and abs(v) > eps:
            last = i
    return values[: last + 1] if last >= 0 else []


def curve_points_from_blocks(data: bytes) -> list[dict[str, float]]:
    current = trim_tail(extract_float_series(data, OFFSETS["curve_current_start"], OFFSETS["curve_current_count"]))
    voltage = trim_tail(extract_float_series(data, OFFSETS["curve_voltage_start"], OFFSETS["curve_voltage_count"]))

    if not current or not voltage:
        return []

    while len(voltage) > 3 and voltage[0] > voltage[1] + 1e-6:
        voltage = voltage[1:]

    n = min(len(current), len(voltage))
    points: list[dict[str, float]] = []
    for i in range(n):
        u = float(voltage[i])
        ia = float(current[i])
        if not (math.isfinite(u) and math.isfinite(ia)):
            continue
        if u < -1 or ia < -1:
            continue
        if u > 2500 or ia > 100:
            continue
        points.append({"u_v": u, "i_a": ia, "p_w": u * ia})
    return points


def derive_metrics_from_curve(points: list[dict[str, float]]) -> dict[str, Any]:
    if not points:
        return {"confidence": "none"}

    points_sorted = sorted(points, key=lambda p: p["u_v"])
    isc = max(points_sorted, key=lambda p: p["i_a"])["i_a"]
    voc = max(points_sorted, key=lambda p: p["u_v"])["u_v"]
    mpp = max(points_sorted, key=lambda p: p["p_w"])
    imp = mpp["i_a"]
    vmp = mpp["u_v"]
    pmax = mpp["p_w"]
    ff = (pmax / (voc * isc)) if voc and isc else None

    return {
        "isc_a": round(isc, 6),
        "voc_v": round(voc, 6),
        "imp_a": round(imp, 6),
        "vmp_v": round(vmp, 6),
        "pmax_w": round(pmax, 6),
        "fill_factor": round(ff, 6) if ff is not None else None,
        "point_count": len(points_sorted),
        "confidence": "medium",
    }


def normalize_filename_only(path_or_name: str | Path) -> str:
    name = os.path.basename(str(path_or_name)).strip()
    return name.lower()


def read_xls_master(xls_path: str | Path) -> dict[str, dict[str, Any]]:
    if pd is None:
        raise RuntimeError("pandas is required to read XLS master file")

    xls_path = Path(xls_path)
    df = pd.read_excel(xls_path)

    # Some exports include a repeated header row as row 0 data.
    # Normalize columns first, then drop rows that look like header echoes.
    normalized_cols = []
    for c in df.columns:
        c_str = str(c).strip()
        normalized_cols.append(XLS_COLUMN_MAP.get(c_str, c_str))
    df.columns = normalized_cols

    # Drop rows with empty filename
    if "filename" not in df.columns:
        raise RuntimeError("Could not find filename column in XLS")

    df = df[df["filename"].notna()].copy()
    df["filename"] = df["filename"].astype(str).str.strip()

    # Remove header echo rows like "Filename"
    df = df[df["filename"].str.lower() != "filename"].copy()

    records: dict[str, dict[str, Any]] = {}
    for _, row in df.iterrows():
        record = {}
        for col in df.columns:
            val = row[col]
            if pd.isna(val):
                continue
            if hasattr(val, "item"):
                try:
                    val = val.item()
                except Exception:
                    pass
            record[col] = val
        key = normalize_filename_only(record["filename"])
        records[key] = record
    return records


def parse_sui_bytes(data: bytes, file_name: str, master_row: dict[str, Any] | None = None, include_curve: bool = True, include_ascii: bool = False) -> dict[str, Any]:
    if not data.startswith(MAGIC):
        raise SUIParseError(f"{file_name}: unrecognized header")

    result: dict[str, Any] = {
        "file_name": file_name,
        "file_size": len(data),
        "format": "UIKenn0File",
        "parser_version": "v4",
        "variant": {
            "family": "PVPMdisp_4_302_fixed_2957",
            "dataset": "hamadia.zip + hamadya.xls",
        },
        "format_version": read_lp_string(data, OFFSETS["format_version"], 20),
        "device": {
            "model_serial": read_lp_string(data, OFFSETS["device_model_serial"], 40),
            "calibration_date": read_lp_string(data, OFFSETS["device_calibration_date"], 20),
        },
        "sensor": {
            "name": read_lp_string(data, OFFSETS["sensor_name"], 60),
            "type": read_lp_string(data, OFFSETS["sensor_type"], 20),
            "calibration_date": read_lp_string(data, OFFSETS["sensor_calibration_date"], 20),
        },
        "module": {
            "manufacturer": read_lp_string(data, OFFSETS["module_manufacturer"], 40),
            "part_number": read_lp_string(data, OFFSETS["module_part_number"], 60),
            "technology": read_lp_string(data, OFFSETS["module_technology"], 20),
            "imp_stc": f32(data, OFFSETS["module_imp_stc"]),
            "voc_stc": f32(data, OFFSETS["module_voc_stc"]),
            "isc_stc": f32(data, OFFSETS["module_isc_stc"]),
            "vmp_stc": f32(data, OFFSETS["module_vmp_stc"]),
            "pmax_stc": f32(data, OFFSETS["module_pmax_stc"]),
            "temp_coeff_current": f32(data, OFFSETS["module_temp_coeff_current"]),
            "temp_coeff_voltage": f32(data, OFFSETS["module_temp_coeff_voltage"]),
            "temp_coeff_power": f32(data, OFFSETS["module_temp_coeff_power"]),
        },
        "measurement": {
            "date_short": read_lp_string(data, OFFSETS["label_date_short"], 20),
            "time": read_lp_string(data, OFFSETS["label_time"], 20),
            "timestamp_text": read_lp_string(data, OFFSETS["label_timestamp_text"], 40),
        },
        "labels": {
            "site_name": read_lp_string(data, OFFSETS["label_site_name"], 60),
            "string_name": read_lp_string(data, OFFSETS["label_string_name"], 60),
        },
    }

    if include_curve:
        points = curve_points_from_blocks(data)
        result["curve"] = {
            "point_count": len(points),
            "points": points,
        }
        result["derived_from_curve"] = derive_metrics_from_curve(points)

    if include_ascii:
        result["ascii_strings"] = ascii_strings(data)

    if master_row:
        result["xls_master"] = master_row
        result["validated_metrics"] = {
            "t_sens": master_row.get("t_sens"),
            "t_mod": master_row.get("t_mod"),
            "e_eff": master_row.get("e_eff"),
            "isc": master_row.get("isc"),
            "uoc": master_row.get("uoc"),
            "ipmax": master_row.get("ipmax"),
            "upmax": master_row.get("upmax"),
            "pmax": master_row.get("pmax"),
            "isc_0": master_row.get("isc_0"),
            "uoc_0": master_row.get("uoc_0"),
            "ipmax_0": master_row.get("ipmax_0"),
            "upmax_0": master_row.get("upmax_0"),
            "ppk": master_row.get("ppk"),
            "fill_factor": master_row.get("fill_factor"),
            "rs": master_row.get("rs"),
            "rp": master_row.get("rp"),
        }
        result["validated_labels"] = {
            "device_number": master_row.get("device_number"),
            "sensor_number": master_row.get("sensor_number"),
            "module": master_row.get("module"),
            "customer": master_row.get("customer"),
            "plant": master_row.get("plant"),
            "part": master_row.get("part"),
            "datetime_text": master_row.get("datetime_text"),
        }
        result["confidence"] = {
            "metadata_from_sui": "high",
            "curve_from_sui": "medium",
            "metrics_from_master_xls": "high",
            "overall": "high",
        }
    else:
        result["confidence"] = {
            "metadata_from_sui": "high",
            "curve_from_sui": "medium",
            "metrics_from_master_xls": "none",
            "overall": "medium",
        }

    return result


def parse_sui_file(path: str | Path, master_index: dict[str, dict[str, Any]] | None = None, include_curve: bool = True, include_ascii: bool = False) -> dict[str, Any]:
    path = Path(path)
    data = path.read_bytes()
    master_row = None
    if master_index:
        master_row = master_index.get(normalize_filename_only(path.name))
    return parse_sui_bytes(data, path.name, master_row=master_row, include_curve=include_curve, include_ascii=include_ascii)


def iter_sui_from_inputs(inputs: list[str | Path]) -> list[tuple[str, bytes]]:
    items: list[tuple[str, bytes]] = []
    for inp in inputs:
        p = Path(inp)
        if p.is_dir():
            for f in sorted(p.rglob("*.SUI")):
                items.append((f.name, f.read_bytes()))
        elif p.suffix.lower() == ".zip":
            with zipfile.ZipFile(p, "r") as zf:
                for name in sorted(zf.namelist()):
                    if name.lower().endswith(".sui"):
                        items.append((os.path.basename(name), zf.read(name)))
        elif p.suffix.lower() == ".sui":
            items.append((p.name, p.read_bytes()))
    return items


def parse_many(inputs: list[str | Path], master_index: dict[str, dict[str, Any]] | None = None, include_curve: bool = True, include_ascii: bool = False) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for file_name, data in iter_sui_from_inputs(inputs):
        master_row = master_index.get(normalize_filename_only(file_name)) if master_index else None
        results.append(parse_sui_bytes(data, file_name, master_row=master_row, include_curve=include_curve, include_ascii=include_ascii))
    return results


def write_single_excel(parsed: dict[str, Any], output_path: str | Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["field", "value"])

    rows = [
        ("file_name", parsed.get("file_name")),
        ("format_version", parsed.get("format_version")),
        ("device_model_serial", parsed.get("device", {}).get("model_serial")),
        ("device_calibration_date", parsed.get("device", {}).get("calibration_date")),
        ("sensor_name", parsed.get("sensor", {}).get("name")),
        ("sensor_type", parsed.get("sensor", {}).get("type")),
        ("sensor_calibration_date", parsed.get("sensor", {}).get("calibration_date")),
        ("module_manufacturer", parsed.get("module", {}).get("manufacturer")),
        ("module_part_number", parsed.get("module", {}).get("part_number")),
        ("module_technology", parsed.get("module", {}).get("technology")),
        ("measurement_date_short", parsed.get("measurement", {}).get("date_short")),
        ("measurement_time", parsed.get("measurement", {}).get("time")),
        ("measurement_timestamp_text", parsed.get("measurement", {}).get("timestamp_text")),
        ("site_name", parsed.get("labels", {}).get("site_name")),
        ("string_name", parsed.get("labels", {}).get("string_name")),
    ]
    for key, value in parsed.get("validated_metrics", {}).items():
        rows.append((f"validated_{key}", value))
    for key, value in parsed.get("validated_labels", {}).items():
        rows.append((f"validated_{key}", value))
    for key, value in parsed.get("derived_from_curve", {}).items():
        rows.append((f"derived_{key}", value))

    for row in rows:
        ws.append(list(row))

    ws2 = wb.create_sheet("Curve")
    ws2.append(["u_v", "i_a", "p_w"])
    for p in parsed.get("curve", {}).get("points", []):
        ws2.append([p.get("u_v"), p.get("i_a"), p.get("p_w")])

    ws3 = wb.create_sheet("Raw_JSON")
    ws3["A1"] = json.dumps(parsed, ensure_ascii=False, indent=2)

    wb.save(output_path)


def write_batch_excel(items: list[dict[str, Any]], output_path: str | Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Batch"
    headers = [
        "file_name",
        "device_model_serial",
        "site_name_sui",
        "string_name_sui",
        "timestamp_text_sui",
        "plant_xls",
        "part_xls",
        "datetime_text_xls",
        "device_number_xls",
        "sensor_number_xls",
        "module_xls",
        "t_sens",
        "t_mod",
        "e_eff",
        "isc",
        "uoc",
        "ipmax",
        "upmax",
        "pmax",
        "isc_0",
        "uoc_0",
        "ipmax_0",
        "upmax_0",
        "ppk",
        "fill_factor",
        "rs",
        "rp",
        "derived_pmax_w",
        "curve_point_count",
        "overall_confidence",
    ]
    ws.append(headers)

    for item in items:
        ws.append([
            item.get("file_name"),
            item.get("device", {}).get("model_serial"),
            item.get("labels", {}).get("site_name"),
            item.get("labels", {}).get("string_name"),
            item.get("measurement", {}).get("timestamp_text"),
            item.get("validated_labels", {}).get("plant"),
            item.get("validated_labels", {}).get("part"),
            item.get("validated_labels", {}).get("datetime_text"),
            item.get("validated_labels", {}).get("device_number"),
            item.get("validated_labels", {}).get("sensor_number"),
            item.get("validated_labels", {}).get("module"),
            item.get("validated_metrics", {}).get("t_sens"),
            item.get("validated_metrics", {}).get("t_mod"),
            item.get("validated_metrics", {}).get("e_eff"),
            item.get("validated_metrics", {}).get("isc"),
            item.get("validated_metrics", {}).get("uoc"),
            item.get("validated_metrics", {}).get("ipmax"),
            item.get("validated_metrics", {}).get("upmax"),
            item.get("validated_metrics", {}).get("pmax"),
            item.get("validated_metrics", {}).get("isc_0"),
            item.get("validated_metrics", {}).get("uoc_0"),
            item.get("validated_metrics", {}).get("ipmax_0"),
            item.get("validated_metrics", {}).get("upmax_0"),
            item.get("validated_metrics", {}).get("ppk"),
            item.get("validated_metrics", {}).get("fill_factor"),
            item.get("validated_metrics", {}).get("rs"),
            item.get("validated_metrics", {}).get("rp"),
            item.get("derived_from_curve", {}).get("pmax_w"),
            item.get("curve", {}).get("point_count"),
            item.get("confidence", {}).get("overall"),
        ])

    wb.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="PVPM SUI parser v4 using Hamadya master XLS for validated output")
    p.add_argument("inputs", nargs="+", help="One or more .SUI files, directories, or .zip archives")
    p.add_argument("--master-xls", help="Path to the master XLS export used to validate and enrich parsed SUI data")
    p.add_argument("--batch", action="store_true", help="Parse all matching SUI files from all inputs")
    p.add_argument("--output-format", choices=["json", "excel"], default="json")
    p.add_argument("-o", "--output", help="Output file path")
    p.add_argument("--ascii", action="store_true", help="Include printable ASCII strings from the SUI binary")
    p.add_argument("--no-curve", action="store_true", help="Skip curve extraction")
    return p


def main() -> None:
    args = build_parser().parse_args()
    include_curve = not args.no_curve

    master_index = None
    if args.master_xls:
        master_index = read_xls_master(args.master_xls)

    if args.batch:
        results = parse_many(args.inputs, master_index=master_index, include_curve=include_curve, include_ascii=args.ascii)
        if args.output_format == "json":
            text = json.dumps(results, ensure_ascii=False, indent=2)
            if args.output:
                Path(args.output).write_text(text, encoding="utf-8")
            else:
                print(text)
        else:
            out = args.output or "pvpm_batch_v4.xlsx"
            write_batch_excel(results, out)
            print(out)
        return

    # Single-file mode: first parseable input only
    input_path = Path(args.inputs[0])
    if input_path.suffix.lower() == ".zip":
        items = iter_sui_from_inputs([input_path])
        if not items:
            raise RuntimeError("No .SUI files found inside ZIP")
        file_name, data = items[0]
        master_row = master_index.get(normalize_filename_only(file_name)) if master_index else None
        result = parse_sui_bytes(data, file_name, master_row=master_row, include_curve=include_curve, include_ascii=args.ascii)
    else:
        result = parse_sui_file(input_path, master_index=master_index, include_curve=include_curve, include_ascii=args.ascii)

    if args.output_format == "json":
        text = json.dumps(result, ensure_ascii=False, indent=2)
        if args.output:
            Path(args.output).write_text(text, encoding="utf-8")
        else:
            print(text)
    else:
        out = args.output or f"{Path(result['file_name']).stem}.parsed.v4.xlsx"
        write_single_excel(result, out)
        print(out)


if __name__ == "__main__":
    main()
