
from __future__ import annotations

import argparse
import json
import math
import re
import struct
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None


class SUIParseError(Exception):
    pass


def f32(data: bytes, offset: int) -> float:
    return struct.unpack_from("<f", data, offset)[0]


def read_lp_string(data: bytes, offset: int, max_len: int = 120, skip_nuls: bool = True) -> tuple[str, int]:
    i = offset
    if skip_nuls:
        while i < len(data) and data[i] == 0:
            i += 1
    if i >= len(data):
        raise SUIParseError(f"Offset out of range: {offset}")

    length = data[i]
    if length <= 0 or length > max_len or i + 1 + length > len(data):
        raise SUIParseError(f"Invalid LP string length at {i}: {length}")

    raw = data[i + 1 : i + 1 + length]
    return raw.decode("latin1", errors="replace").strip(), i + 1 + length


def safe_lp(data: bytes, offset: int) -> Optional[str]:
    try:
        value, _ = read_lp_string(data, offset)
        return value
    except Exception:
        return None


def extract_curve(data: bytes, start: int, count: int) -> list[float]:
    out: list[float] = []
    for i in range(count):
        off = start + i * 4
        if off + 4 > len(data):
            break
        try:
            out.append(f32(data, off))
        except Exception:
            break
    return out


def trim_tail(values: list[float]) -> list[float]:
    last = -1
    for i, v in enumerate(values):
        if math.isfinite(v) and abs(v) > 1e-12:
            last = i
    return values[: last + 1] if last >= 0 else []


def ascii_strings(data: bytes, min_len: int = 4) -> list[str]:
    found = re.findall(rb"[ -~]{%d,}" % min_len, data)
    return [s.decode("latin1", errors="replace") for s in found]


# Sample-derived offsets from uploaded PVPMdisp/Curvealyzer files.
OFFSETS = {
    "format_version": 153,
    "device": 160,
    "device_calibration_date": 175,
    "sensor_name": 233,
    "sensor_type": 247,
    "sensor_calibration_date": 280,
    "module_manufacturer": 300,
    "module_type": 326,
    "module_technology": 375,
    "module_imp_stc": 388,
    "module_voc_stc": 392,
    "module_isc_stc": 396,
    "module_vmp_stc": 400,
    "module_pmax_stc": 404,
    "module_tc_current": 416,
    "module_tc_voltage": 420,
    "module_tc_power": 424,
    "curve_current_start": 503,
    "curve_voltage_start": 1475,
    "label_date_short": 2477,
    "label_time": 2486,
    "label_code_1": 2546,
    "label_code_2": 2597,
    "label_timestamp_text": 2648,
}


def normalize_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def best_timestamp(date_short: Optional[str], time_text: Optional[str], timestamp_text: Optional[str]) -> Optional[str]:
    patterns = [
        ("%d.%m.%Y  %H:%M:%S", timestamp_text),
        ("%d.%m.%Y %H:%M:%S", timestamp_text),
        ("%d.%m.%y %H:%M:%S", f"{date_short} {time_text}" if date_short and time_text else None),
        ("%d.%m.%Y %H:%M:%S", f"{date_short} {time_text}" if date_short and time_text else None),
    ]
    for fmt, value in patterns:
        if not value:
            continue
        try:
            dt = datetime.strptime(value.strip(), fmt)
            return dt.isoformat()
        except Exception:
            pass
    return None


def derive_from_curve(points: list[dict[str, float]]) -> dict[str, Optional[float]]:
    if not points:
        return {
            "isc_a": None,
            "voc_v": None,
            "imp_a": None,
            "vmp_v": None,
            "pmax_w": None,
            "fill_factor": None,
        }

    # sort by voltage ascending, guard duplicates
    pts = sorted(points, key=lambda x: x["u_v"])
    isc = pts[0]["i_a"]
    # last point where current approaches 0
    voc = None
    for p in reversed(pts):
        if abs(p["i_a"]) < 0.2:
            voc = p["u_v"]
            break
    if voc is None:
        voc = pts[-1]["u_v"]

    pmax_point = max(pts, key=lambda x: x["p_w"])
    pmax = pmax_point["p_w"]
    vmp = pmax_point["u_v"]
    imp = pmax_point["i_a"]

    ff = None
    denom = voc * isc if voc is not None and isc is not None else None
    if denom and abs(denom) > 1e-9:
        ff = pmax / denom

    return {
        "isc_a": round(isc, 6) if isc is not None else None,
        "voc_v": round(voc, 6) if voc is not None else None,
        "imp_a": round(imp, 6) if imp is not None else None,
        "vmp_v": round(vmp, 6) if vmp is not None else None,
        "pmax_w": round(pmax, 6) if pmax is not None else None,
        "fill_factor": round(ff, 6) if ff is not None else None,
    }


def parse_sui(path: str | Path, include_curve: bool = True, include_ascii_dump: bool = False) -> dict[str, Any]:
    path = Path(path)
    data = path.read_bytes()

    if not data.startswith(b"UIKenn0File"):
        raise SUIParseError("Unrecognized SUI header")

    out: dict[str, Any] = {
        "file_name": path.name,
        "file_size": len(data),
        "format": "UIKenn0File",
        "source": {"sui_path": str(path)},
    }

    version = safe_lp(data, OFFSETS["format_version"])
    if version:
        out["format_version"] = version

    device = safe_lp(data, OFFSETS["device"])
    device_cal = safe_lp(data, OFFSETS["device_calibration_date"])
    out["device"] = {
        "model_serial": device,
        "calibration_date": device_cal,
    }

    sensor_name = safe_lp(data, OFFSETS["sensor_name"])
    sensor_type = safe_lp(data, OFFSETS["sensor_type"])
    sensor_cal = safe_lp(data, OFFSETS["sensor_calibration_date"])
    out["sensor"] = {
        "name": sensor_name,
        "type": sensor_type,
        "calibration_date": sensor_cal,
    }

    module_manufacturer = safe_lp(data, OFFSETS["module_manufacturer"])
    module_type = safe_lp(data, OFFSETS["module_type"])
    module_technology = safe_lp(data, OFFSETS["module_technology"])
    out["module"] = {
        "manufacturer": module_manufacturer,
        "part_number": module_type,
        "technology": module_technology,
        "imp_stc": f32(data, OFFSETS["module_imp_stc"]),
        "voc_stc": f32(data, OFFSETS["module_voc_stc"]),
        "isc_stc": f32(data, OFFSETS["module_isc_stc"]),
        "vmp_stc": f32(data, OFFSETS["module_vmp_stc"]),
        "pmax_stc": f32(data, OFFSETS["module_pmax_stc"]),
        "temp_coeff_current": f32(data, OFFSETS["module_tc_current"]),
        "temp_coeff_voltage": f32(data, OFFSETS["module_tc_voltage"]),
        "temp_coeff_power": f32(data, OFFSETS["module_tc_power"]),
    }

    date_short = safe_lp(data, OFFSETS["label_date_short"])
    time_text = safe_lp(data, OFFSETS["label_time"])
    code1 = safe_lp(data, OFFSETS["label_code_1"])
    code2 = safe_lp(data, OFFSETS["label_code_2"])
    timestamp_text = safe_lp(data, OFFSETS["label_timestamp_text"])
    out["labels"] = {
        "date_short": date_short,
        "time": time_text,
        "code_1": code1,
        "code_2": code2,
        "timestamp_text": timestamp_text,
        "timestamp_iso": best_timestamp(date_short, time_text, timestamp_text),
    }

    if include_curve:
        current = trim_tail(extract_curve(data, OFFSETS["curve_current_start"], 110))
        voltage = trim_tail(extract_curve(data, OFFSETS["curve_voltage_start"], 130))

        if current and voltage:
            while len(voltage) > 5 and voltage[0] > voltage[1] + 1e-6:
                voltage = voltage[1:]

            n = min(len(current), len(voltage))
            points = []
            for i in range(n):
                u = float(voltage[i])
                c = float(current[i])
                points.append({
                    "index": i + 1,
                    "u_v": u,
                    "i_a": c,
                    "p_w": u * c,
                })

            out["curve"] = {
                "point_count": n,
                "points": points,
            }
            out["derived_from_curve"] = derive_from_curve(points)

    if include_ascii_dump:
        out["ascii_strings"] = ascii_strings(data)

    return out


def find_matching_xls(sui_path: Path, explicit_xls: Optional[str] = None) -> Optional[Path]:
    if explicit_xls:
        p = Path(explicit_xls)
        return p if p.exists() else None

    candidates = [
        sui_path.with_suffix(".XLS"),
        sui_path.with_suffix(".xls"),
        sui_path.parent / (sui_path.stem + ".XLS"),
        sui_path.parent / (sui_path.stem + ".xls"),
        sui_path.parent / "data.XLS",
        sui_path.parent / "data.xls",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def parse_xls(xls_path: Path) -> dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for XLS/XLSX parsing in this script.")

    # openpyxl can read many HTML/SpreadsheetML-as-.XLS exports; true BIFF may fail.
    wb = load_workbook(filename=xls_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    flat = [[c for c in row] for row in rows]

    result: dict[str, Any] = {
        "source": {"xls_path": str(xls_path)},
        "summary": {},
        "curve": {"points": []},
        "raw_sheet_name": ws.title,
    }

    # Find key/value pairs in first ~60 rows
    key_map = {
        "device": ["device", "gerät", "pvpm"],
        "sensor": ["sensor"],
        "isc": ["isc"],
        "uoc": ["uoc", "voc"],
        "ipmax": ["ipmax", "imp"],
        "upmax": ["upmax", "vmp"],
        "ppk": ["ppk", "pmax"],
        "t sens": ["t sens", "tsens"],
        "t mod": ["t mod", "tmod"],
        "e eff": ["e eff", "irradiance", "eeff"],
        "date": ["date", "datum"],
        "time": ["time", "uhrzeit"],
    }

    def set_summary(k: str, v: Any):
        if v not in (None, ""):
            result["summary"][k] = v

    for row in flat[:80]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        joined = " | ".join(cells).lower()
        for dest, keys in key_map.items():
            if any(key in joined for key in keys):
                # take last non-empty cell on row as value
                value = None
                for c in reversed(row):
                    if c not in (None, ""):
                        value = c
                        break
                set_summary(dest, value)

    # Detect curve header
    header_idx = None
    for i, row in enumerate(flat):
        normalized = [str(c).strip().lower() if c is not None else "" for c in row]
        if "u in v" in normalized and "i in a" in normalized:
            header_idx = i
            break

    if header_idx is not None:
        headers = [str(c).strip().lower() if c is not None else "" for c in flat[header_idx]]
        try:
            u_idx = headers.index("u in v")
            i_idx = headers.index("i in a")
        except ValueError:
            u_idx = i_idx = None

        p_idx = headers.index("p in w") if "p in w" in headers else None

        if u_idx is not None and i_idx is not None:
            for row in flat[header_idx + 1 :]:
                if max(len(row), u_idx + 1, i_idx + 1) <= max(u_idx, i_idx):
                    continue
                u = row[u_idx] if u_idx < len(row) else None
                i_val = row[i_idx] if i_idx < len(row) else None
                p = row[p_idx] if p_idx is not None and p_idx < len(row) else None
                u_f = normalize_float(u)
                i_f = normalize_float(i_val)
                p_f = normalize_float(p)
                if u_f is None or i_f is None:
                    continue
                if p_f is None:
                    p_f = u_f * i_f
                result["curve"]["points"].append({
                    "index": len(result["curve"]["points"]) + 1,
                    "u_v": u_f,
                    "i_a": i_f,
                    "p_w": p_f,
                })

    result["curve"]["point_count"] = len(result["curve"]["points"])
    if result["curve"]["points"]:
        result["derived_from_curve"] = derive_from_curve(result["curve"]["points"])

    # normalize some common summary values to numeric
    num_fields = ["isc", "uoc", "ipmax", "upmax", "ppk", "t sens", "t mod", "e eff"]
    for f in num_fields:
        if f in result["summary"]:
            result["summary"][f] = normalize_float(result["summary"][f])

    return result


def enrich_with_xls(parsed: dict[str, Any], xls_data: dict[str, Any]) -> dict[str, Any]:
    enriched = json.loads(json.dumps(parsed))  # deep copy via JSON-safe route
    enriched["matching_xls"] = xls_data

    # Prefer explicit XLS summary for measurement figures where present.
    s = xls_data.get("summary", {})
    measurement = {
        "isc_a": s.get("isc"),
        "voc_v": s.get("uoc"),
        "imp_a": s.get("ipmax"),
        "vmp_v": s.get("upmax"),
        "pmax_w": s.get("ppk"),
        "t_sensor_c": s.get("t sens"),
        "t_module_c": s.get("t mod"),
        "irradiance_w_m2": s.get("e eff"),
    }
    if any(v is not None for v in measurement.values()):
        enriched["measurement"] = measurement

    if xls_data.get("curve", {}).get("points"):
        enriched["curve"] = xls_data["curve"]
        enriched["derived_from_curve"] = xls_data.get("derived_from_curve")

    if s.get("device"):
        enriched.setdefault("device", {})["reported_by_xls"] = s.get("device")
    if s.get("sensor"):
        enriched.setdefault("sensor", {})["reported_by_xls"] = s.get("sensor")

    # Best timestamp from XLS if present
    date_value = s.get("date")
    time_value = s.get("time")
    if date_value or time_value:
        enriched.setdefault("labels", {})["xls_date"] = date_value
        enriched.setdefault("labels", {})["xls_time"] = time_value

    return enriched


def flatten_summary(parsed: dict[str, Any]) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    rows.extend([
        ("file_name", parsed.get("file_name")),
        ("file_size", parsed.get("file_size")),
        ("format", parsed.get("format")),
        ("format_version", parsed.get("format_version")),
    ])
    for section in ["device", "sensor", "module", "labels", "measurement", "derived_from_curve"]:
        block = parsed.get(section, {})
        if isinstance(block, dict):
            for k, v in block.items():
                rows.append((f"{section}.{k}", v))
    if "matching_xls" in parsed:
        xls = parsed["matching_xls"]
        for k, v in xls.get("summary", {}).items():
            rows.append((f"xls.summary.{k}", v))
    return rows


def write_excel(parsed: dict[str, Any], out_path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel output.")

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    for k, v in flatten_summary(parsed):
        ws.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])

    # Curve sheet
    curve = parsed.get("curve", {})
    points = curve.get("points", []) if isinstance(curve, dict) else []
    ws2 = wb.create_sheet("Curve")
    ws2.append(["Index", "Voltage (V)", "Current (A)", "Power (W)"])
    for p in points:
        ws2.append([p.get("index"), p.get("u_v"), p.get("i_a"), p.get("p_w")])

    # Raw JSON sheet (chunked to lines)
    ws3 = wb.create_sheet("Raw_JSON")
    raw_json = json.dumps(parsed, ensure_ascii=False, indent=2).splitlines()
    ws3.append(["json"])
    for line in raw_json:
        ws3.append([line])

    wb.save(out_path)


def output_basename(sui_path: Path) -> str:
    return sui_path.stem.replace("  ", "_").replace(" ", "_").replace(":", "-")


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse PVPMdisp/Curvealyzer .SUI files with JSON or Excel output.")
    parser.add_argument("input_file", help="Path to .SUI file")
    parser.add_argument("-o", "--output", help="Output file path")
    parser.add_argument("--output-format", choices=["json", "excel"], default="json", help="Output format")
    parser.add_argument("--xls", help="Optional matching XLS file to merge")
    parser.add_argument("--no-xls", action="store_true", help="Do not auto-merge matching XLS")
    parser.add_argument("--ascii", action="store_true", help="Include ASCII string dump from SUI")
    parser.add_argument("--no-curve", action="store_true", help="Skip SUI curve extraction")
    args = parser.parse_args()

    sui_path = Path(args.input_file)
    parsed = parse_sui(sui_path, include_curve=not args.no_curve, include_ascii_dump=args.ascii)

    if not args.no_xls:
        xls_path = find_matching_xls(sui_path, args.xls)
        if xls_path is not None:
            try:
                xls_data = parse_xls(xls_path)
                parsed = enrich_with_xls(parsed, xls_data)
            except Exception as exc:
                parsed.setdefault("warnings", []).append(f"Failed to parse XLS '{xls_path}': {exc}")

    if args.output:
        out_path = Path(args.output)
    else:
        ext = ".json" if args.output_format == "json" else ".xlsx"
        out_path = sui_path.with_name(output_basename(sui_path) + ext)

    if args.output_format == "json":
        out_path.write_text(json.dumps(parsed, ensure_ascii=False, indent=2), encoding="utf-8")
    else:
        write_excel(parsed, out_path)

    print(str(out_path))


if __name__ == "__main__":
    main()
